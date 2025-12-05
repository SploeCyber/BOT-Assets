import os
import re
import json
import argparse
from collections import OrderedDict
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

# Configuration
INPUT_FOLDER = "Downloads"
ASSETS_FOLDER = "Assets"

def sanitize_filename(name):
    """Removes non-printable characters and slashes."""
    if not name:
        return ""
    # Remove non-printable characters (0-31, 127-159)
    name = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', str(name))
    # Replace slashes
    name = name.replace("/", "-").replace("\\", "-")
    return name.strip()

def to_int_if_possible(value):
    """Converts value to int if it's a whole number, otherwise returns original."""
    if value is None:
        return None
    try:
        f = float(value)
        if f.is_integer():
            return int(f)
        return f
    except (ValueError, TypeError):
        return value

class MergedCellLookup:
    """Helper to quickly resolve values for merged cells."""
    def __init__(self, sheet):
        self.sheet = sheet
        self.lookup = {}
        self._build_lookup()

    def _build_lookup(self):
        for merged in self.sheet.merged_cells.ranges:
            # merged.min_row, merged.min_col is the top-left cell
            val = self.sheet.cell(merged.min_row, merged.min_col).value
            # Map every cell in this range to the top-left value
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    self.lookup[(row, col)] = val

    def get_value(self, row, col):
        """Returns the value from the merged cell if it exists, otherwise the cell's own value."""
        if (row, col) in self.lookup:
             return self.lookup[(row, col)]
        return self.sheet.cell(row=row, column=col).value


def process_card_image(idx, image, ws, merged_lookup, args, output_folder, card_count_so_far, seen_prints):
    """Extracts data and image for a single card."""
    try:
        # The image anchor gives the top-left placement.
        # Note: openpyxl image anchors are 0-indexed for row/col in some versions
        # but typically _from.row is 0-based index. 
        # In the original script: img_row = image.anchor._from.row + 1 (to make it 1-based)
        img_row_idx = image.anchor._from.row # 0-based
        img_row = img_row_idx + 1 # 1-based
        data_row = img_row + 1

        # Check for card Name
        card_name = ws.cell(row=data_row, column=3).value
        if not card_name:
            # print(f"  Skipping image at row {img_row}: No card name found at row {data_row}")
            return None

        # --- Image Processing ---
        img_obj = None
        if not args.no_image:
             img_bytes = image._data()
             if isinstance(img_bytes, bytes):
                 img_obj = Image.open(BytesIO(img_bytes))
             else:
                 print(f"  Warning: Invalid image data at row {img_row}")
                 if not args.no_image: return None # Strictly require image unless flag

        # --- Data Extraction ---
        card_temp = {}
        
        # Columns 3 to 14 (C to N)
        keys = [ws.cell(row=img_row, column=c).value for c in range(3, 15)]
        vals = [ws.cell(row=data_row, column=c).value for c in range(3, 15)]

        for k, v in zip(keys, vals):
            if k:
                card_temp[str(k)] = to_int_if_possible(v)

        # --- Details Extraction ---
        details = {}
        for r in range(data_row + 1, data_row + 4):
            # key at col 3, value at col 4
            # Use lookup for merged cells
            key_val = merged_lookup.get_value(r, 3)
            val_val = merged_lookup.get_value(r, 4)
            
            if key_val:
                details[str(key_val)] = to_int_if_possible(val_val)
        
        if details:
            card_temp["Details"] = details

        # --- Color Extraction ---
        bg_color = ws.cell(row=data_row, column=9).fill.fgColor.rgb
        # Typical transparent/empty is 00000000 or None
        if bg_color and bg_color != "00000000" and len(str(bg_color)) >= 6:
             # ARGB -> get last 6 chars usually
             card_temp["Color"] = f"#{str(bg_color)[-6:]}"
        else:
             card_temp["Color"] = None

        # --- Cleanup / Logic ---
        if "Ex" in card_temp and card_temp["Ex"] is None:
            del card_temp["Ex"]

        if card_temp.get("Type") == "Construct":
            card_temp.pop("Color", None)

        if card_temp.get("Type") == "Magic":
            card_temp.pop("Power", None)
            card_temp.pop("Color", None)
            if "Cost" in card_temp:
                 card_temp["SubType"] = card_temp.pop("Cost")

        card_temp.pop("Drop Rate", None)

        # Define Filename & Duplicate Logic
        print_code = str(card_temp.get("Print", f"card_{card_count_so_far+1}")).strip()
        rarity = str(card_temp.get("Rare", "")).strip()
        
        original_print = print_code
        original_rare = rarity

        if print_code not in seen_prints:
            seen_prints[print_code] = [original_rare]
            
            # Normal Filename
            if original_rare:
                file_name = f"{original_print}-{original_rare}.png"
            else:
                file_name = f"{original_print}.png"
        else:
            # It's a duplicate Print
            seen_prints[print_code].append(original_rare)
            count = len(seen_prints[print_code])
            
            # Modify Print: {Print}-{Count}
            new_print = f"{original_print}-{count}"
            card_temp["Print"] = new_print
            
            # Check if Rare is also a duplicate for this Print
            # We check if original_rare appeared in the list BEFORE this current insertion
            previous_rares = seen_prints[print_code][:-1]
            
            if original_rare in previous_rares:
                new_rare = f"{original_rare}-{count}"
                card_temp["Rare"] = new_rare
            else:
                new_rare = original_rare
            
            # Modify Filename: {NewPrint}-{NewRare}.png
            if new_rare:
                file_name = f"{new_print}-{new_rare}.png"
            else:
                file_name = f"{new_print}.png"

        # Sanitize Filename
        file_name = sanitize_filename(file_name)

        # Old Logic (to be removed)
        # Define Filename
        print_code = str(card_temp.get("Print", f"card_{card_count_so_far+1}")).strip()
        rarity = str(card_temp.get("Rare", "")).strip()
        name_part = f"{print_code}-{rarity}" if rarity else print_code
        safe_name = name_part.replace("/", "-").replace("\\", "-")
        file_name = f"{safe_name}.png"
        
        # Save Image
        if img_obj:
            img_path = os.path.join(output_folder, file_name)
            img_obj.save(img_path)
            
        card_temp["ImagePath"] = file_name

        # Post-Processing Logic
        details_text = json.dumps(card_temp.get("Details", {}), ensure_ascii=False)
        if "เมื่อการ์ดใบนี้ถูกหงายจากการโจมตี" in details_text:
            card_temp["Type"] = "Life"
        
        if str(card_temp.get("Type", "")).strip().lower() == "life":
            # Remove empty fields for Life cards
            card_temp = {k: v for k, v in card_temp.items() if v not in (None, "", " ")}

        # Reorder Fields
        dataset_card = OrderedDict()
        dataset_card["ImagePath"] = card_temp.pop("ImagePath", None)
        
        subtype_val = card_temp.pop("SubType", None)

        for k, v in card_temp.items():
            dataset_card[k] = v
            # Insert SubType immediately after Type
            if k == "Type" and subtype_val is not None:
                dataset_card["SubType"] = subtype_val

        return dataset_card, safe_name

    except Exception as e:
        print(f"  Error processing row {img_row}: {e}")
        return None, None


def generate_assets_index(assets_folder):
    """Scans the Assets folder and generates a root-level index JSON."""
    print(f"\nGenerating asset index...")
    asset_list = []
    
    if os.path.exists(assets_folder):
        for folder_name in sorted(os.listdir(assets_folder)):
            folder_path = os.path.join(assets_folder, folder_name)
            if os.path.isdir(folder_path):
                dataset_path = os.path.join(folder_path, "dataset.json")
                if os.path.exists(dataset_path):
                    # Create the entry
                    # Use forward slashes for path consistency
                    relative_path = f"{ASSETS_FOLDER}/{folder_name}/dataset.json"
                    asset_list.append({
                        "name": folder_name,
                        "path": relative_path
                    })
    
    output_path = "asset-map.json"
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(asset_list, f, ensure_ascii=False, indent=2)
        print(f"✅ Asset map saved to: {output_path}")
        print(f"📦 Total Asset Packs: {len(asset_list)}")
    except Exception as e:
        print(f"❌ Error saving asset map: {e}")


def main():
    parser = argparse.ArgumentParser(description="Generate JSON from Excel with card data.")
    parser.add_argument("--no-image", action="store_true", help="Skip image extraction and saving.")
    args = parser.parse_args()

    # Find XLSX files
    xlsx_files = []
    if os.path.exists(INPUT_FOLDER):
        for root, dirs, files in os.walk(INPUT_FOLDER):
            for file in files:
                if file.endswith(".xlsx") and not file.startswith("~$"):
                    xlsx_files.append(os.path.join(root, file))

    if not xlsx_files:
        print(f"ไม่พบไฟล์ .xlsx ในโฟลเดอร์ {INPUT_FOLDER}")
        return

    for xlsx_path in xlsx_files:
        print(f"\nProcessing {os.path.basename(xlsx_path)}...")
        
        file_name = os.path.basename(xlsx_path)
        folder_name = os.path.splitext(file_name)[0]
        # Sanitize folder name
        folder_name = sanitize_filename(folder_name)
        output_folder = os.path.join(ASSETS_FOLDER, folder_name)
        os.makedirs(output_folder, exist_ok=True)

        cards = []
        try:
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            
            # 1. Validation
            if not hasattr(ws, '_images'):
                print(f"  No images found (metadata missing).")
                continue
            
            # 2. Sort images by row position
            sorted_images = sorted(ws._images, key=lambda img: img.anchor._from.row)
            print(f"  Found {len(sorted_images)} images. Analyzing...")

            # 3. Build Merged Cell Lookup
            merged_lookup = MergedCellLookup(ws)
            
            # Track seen prints for duplicate handling
            seen_prints = {}

            # 4. Process each image as a card entry
            for i, image in enumerate(sorted_images):
                card_data, card_name = process_card_image(
                    i, image, ws, merged_lookup, args, output_folder, len(cards), seen_prints
                )
                if card_data:
                    cards.append(card_data)
                    print(f"  Processed: {card_name}")
            
            # 5. Save JSON
            json_path = os.path.join(output_folder, "dataset.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(cards, f, ensure_ascii=False, indent=2)

            print(f"✅ Success! Saved to: {json_path}")
            print(f"📦 Total Cards: {len(cards)}")

        except Exception as e:
            print(f"❌ Critical Error reading {xlsx_path}: {e}")

    # Generate the root index file after processing all sheets
    generate_assets_index(ASSETS_FOLDER)

if __name__ == "__main__":
    main()