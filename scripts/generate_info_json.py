#!/usr/bin/env python3
"""
Generate card dataset JSON files from Excel spreadsheets.

Extracts card images and metadata from BOT card game Excel files,
generates structured card-data.json files, and creates asset index.

Usage:
    python generate_info_json.py                  # Process all XLSX files
    python generate_info_json.py --no-image       # Skip image extraction
"""

import os
import re
import json
import logging
import argparse
import hashlib
from io import BytesIO
from collections import OrderedDict
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl import load_workbook
from PIL import Image

from card_overrides import apply_card_overrides

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Configuration
MAX_IMAGE_THREADS: int = 8
MAX_FILE_THREADS: int = 4
CARD_DATA_FILENAME: str = "card-data.json"
ALL_CARDS_FILENAME: str = "all-cards.json"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
INPUT_FOLDER = os.environ.get("INPUT_FOLDER", os.path.join(BASE_DIR, "Downloads"))
ASSETS_FOLDER = os.path.join(BASE_DIR, "assets")

def sanitize_filename(name):
    """Removes non-printable characters and slashes."""
    if not name:
        return ""
    # Remove non-printable characters (0-31, 127-159)
    name = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', str(name))
    # Replace slashes
    name = name.replace("/", "-").replace("\\", "-")
    return name.strip()


CARD_DATA_FILENAME = "card-data.json"


def to_int_if_possible(value):
    """Converts value to int if it's a whole number, otherwise returns original."""
    if value is None:
        return None
    try:
        f = float(value)
        if f.is_integer():
            return int(f)
        return f
    except (ValueError, TypeError):
        return value

class MergedCellLookup:
    """Helper to quickly resolve values for merged cells."""
    def __init__(self, sheet):
        self.sheet = sheet
        self.lookup = {}
        self._build_lookup()

    def _build_lookup(self):
        for merged in self.sheet.merged_cells.ranges:
            # merged.min_row, merged.min_col is the top-left cell
            val = self.sheet.cell(merged.min_row, merged.min_col).value
            # Map every cell in this range to the top-left value
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    self.lookup[(row, col)] = val

    def get_value(self, row, col):
        """Returns the value from the merged cell if it exists, otherwise the cell's own value."""
        if (row, col) in self.lookup:
             return self.lookup[(row, col)]
        return self.sheet.cell(row=row, column=col).value


def process_card_image(idx, image, ws, merged_lookup, args, output_folder, card_count_so_far, seen_prints):
    """Extracts data and image for a single card."""
    try:
        # The image anchor gives the top-left placement.
        # Note: openpyxl image anchors are 0-indexed for row/col in some versions
        # but typically _from.row is 0-based index. 
        # In the original script: img_row = image.anchor._from.row + 1 (to make it 1-based)
        img_row_idx = image.anchor._from.row # 0-based
        img_row = img_row_idx + 1 # 1-based
        data_row = img_row + 1

        # Check for card Name
        card_name = ws.cell(row=data_row, column=3).value
        if not card_name:
            # print(f"  Skipping image at row {img_row}: No card name found at row {data_row}")
            return None

        # --- Image Processing ---
        img_obj = None
        if not args.no_image:
             img_bytes = image._data()
             if isinstance(img_bytes, bytes):
                 img_obj = Image.open(BytesIO(img_bytes))
             else:
                 logger.warning(f"Invalid image data at row {img_row}")
                 if not args.no_image:
                     return None  # Strictly require image unless flag

        # --- Data Extraction ---
        card_temp = {}
        
        # Columns 3 to 14 (C to N)
        keys = [ws.cell(row=img_row, column=c).value for c in range(3, 15)]
        vals = [ws.cell(row=data_row, column=c).value for c in range(3, 15)]

        for k, v in zip(keys, vals):
            if k:
                card_temp[str(k)] = to_int_if_possible(v)

        # --- Details Extraction ---
        details = {}
        for r in range(data_row + 1, data_row + 4):
            # key at col 3, value at col 4
            # Use lookup for merged cells
            key_val = merged_lookup.get_value(r, 3)
            val_val = merged_lookup.get_value(r, 4)
            
            if key_val:
                details[str(key_val)] = to_int_if_possible(val_val)
        
        if details:
            card_temp["Details"] = details

        # --- Color Extraction ---
        bg_color = ws.cell(row=data_row, column=9).fill.fgColor.rgb
        # Typical transparent/empty is 00000000 or None
        if bg_color and bg_color != "00000000" and len(str(bg_color)) >= 6:
             # ARGB -> get last 6 chars usually
             card_temp["Color"] = f"#{str(bg_color)[-6:]}"
        else:
             card_temp["Color"] = None

        # --- Cleanup / Logic ---
        if "Ex" in card_temp and card_temp["Ex"] is None:
            del card_temp["Ex"]

        if card_temp.get("Type") == "Construct":
            card_temp.pop("Color", None)

        if card_temp.get("Type") == "Magic":
            card_temp.pop("Power", None)
            card_temp.pop("Color", None)
            if "Cost" in card_temp:
                 card_temp["SubType"] = card_temp.pop("Cost")

        card_temp.pop("Drop Rate", None)

        # Post-Processing Logic
        details_text = json.dumps(card_temp.get("Details", {}), ensure_ascii=False)
        if "เมื่อการ์ดใบนี้ถูกหงายจากการโจมตี" in details_text:
            card_temp["Type"] = "Life"
        
        if str(card_temp.get("Type", "")).strip().lower() == "life":
            # Remove empty fields for Life cards
            card_temp = {k: v for k, v in card_temp.items() if v not in (None, "", " ")}

        # Apply card overrides BEFORE filename determination
        # so overridden values (Rare, Print, etc.) affect filenames
        original_print = str(card_temp.get("Print", "")).strip()
        card_temp = apply_card_overrides(card_temp, original_print)

        # Define Filename & Duplicate Logic (uses overridden values)
        print_code = str(card_temp.get("Print", f"card_{card_count_so_far+1}")).strip()
        rarity = str(card_temp.get("Rare", "")).strip()
        
        original_print = print_code
        original_rare = rarity

        if print_code not in seen_prints:
            seen_prints[print_code] = [original_rare]
            
            # Normal Filename
            if original_rare:
                file_name = f"{original_print}-{original_rare}.png"
            else:
                file_name = f"{original_print}.png"
        else:
            # It's a duplicate Print
            seen_prints[print_code].append(original_rare)
            count = len(seen_prints[print_code])
            
            # Modify Print: {Print}-{Count}
            new_print = f"{original_print}-{count}"
            card_temp["Print"] = new_print
            
            # Check if Rare is also a duplicate for this Print
            previous_rares = seen_prints[print_code][:-1]
            
            if original_rare in previous_rares:
                new_rare = f"{original_rare}-{count}"
                card_temp["Rare"] = new_rare
            else:
                new_rare = original_rare
            
            # Modify Filename: {NewPrint}-{NewRare}.png
            if new_rare:
                file_name = f"{new_print}-{new_rare}.png"
            else:
                file_name = f"{new_print}.png"

        # Sanitize Filename
        file_name = sanitize_filename(file_name)
        safe_name = os.path.splitext(file_name)[0]
        
        # Prepare image for deferred (threaded) saving
        img_save_task = None
        if img_obj:
            img_path = os.path.join(output_folder, file_name)
            img_save_task = (img_obj, img_path)
            
        card_temp["ImagePath"] = file_name

        # Reorder Fields
        dataset_card = OrderedDict()
        dataset_card["ImagePath"] = card_temp.pop("ImagePath", None)
        
        subtype_val = card_temp.pop("SubType", None)

        for k, v in card_temp.items():
            dataset_card[k] = v
            # Insert SubType immediately after Type
            if k == "Type" and subtype_val is not None:
                dataset_card["SubType"] = subtype_val

        return dataset_card, safe_name, img_save_task

    except Exception as e:
        logger.error(f"Error processing row {img_row}: {e}")
        return None, None, None


def generate_assets_index(assets_folder: str) -> None:
    """Scans the assets folder and generates a root-level index JSON."""
    logger.info("\nGenerating asset index...")
    asset_list = []

    # Add all-cards.json as the first entry (merged dataset)
    all_cards_path = os.path.join(assets_folder, ALL_CARDS_FILENAME)
    if os.path.exists(all_cards_path):
        asset_list.append({
            "name": "All Cards",
            "path": f"assets/{ALL_CARDS_FILENAME}"
        })

    if os.path.exists(assets_folder):
        for folder_name in sorted(os.listdir(assets_folder)):
            folder_path = os.path.join(assets_folder, folder_name)
            if os.path.isdir(folder_path):
                dataset_path = os.path.join(folder_path, CARD_DATA_FILENAME)
                if os.path.exists(dataset_path):
                    # Create the entry
                    # Use forward slashes for path consistency
                    relative_path = f"assets/{folder_name}/{CARD_DATA_FILENAME}"
                    asset_list.append({
                        "name": folder_name,
                        "path": relative_path
                    })

    output_path = os.path.join(BASE_DIR, "asset-map.json")
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(asset_list, f, ensure_ascii=False, indent=2)
        logger.info(f"✅ Asset map saved to: {output_path}")
        logger.info(f"📦 Total Asset Packs: {len(asset_list)}")
    except Exception as e:
        logger.error(f"❌ Error saving asset map: {e}")


def save_image_if_changed(img_obj: Image.Image, img_path: str) -> bool:
    """Save image only if pixel content differs from existing file. Thread-safe."""
    try:
        new_pixel_hash = hashlib.md5(img_obj.tobytes()).hexdigest()

        if os.path.exists(img_path):
            try:
                existing_img = Image.open(img_path)
                existing_pixel_hash = hashlib.md5(existing_img.tobytes()).hexdigest()
                existing_img.close()
                if existing_pixel_hash == new_pixel_hash:
                    return False  # No change
            except Exception:
                pass  # Can't read existing, overwrite

        img_obj.save(img_path, optimize=True)
        return True  # Saved
    except Exception as e:
        logger.warning(f"⚠️  Error saving {img_path}: {e}")
        return False


def process_single_xlsx(xlsx_path: str, args) -> int:
    """Process a single XLSX file: extract card data and save images with threading."""
    logger.info(f"\nProcessing {os.path.basename(xlsx_path)}...")

    file_name = os.path.basename(xlsx_path)
    folder_name = os.path.splitext(file_name)[0]
    folder_name = sanitize_filename(folder_name)
    output_folder = os.path.join(ASSETS_FOLDER, folder_name)
    os.makedirs(output_folder, exist_ok=True)

    cards = []
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # 1. Validation
        if not hasattr(ws, '_images'):
            logger.warning(f"No images found (metadata missing).")
            return 0

        # 2. Sort images by row position
        sorted_images = sorted(ws._images, key=lambda img: img.anchor._from.row)
        logger.info(f"Found {len(sorted_images)} images. Analyzing...")

        # 3. Build Merged Cell Lookup
        merged_lookup = MergedCellLookup(ws)

        # Track seen prints for duplicate handling
        seen_prints: Dict[str, List[Optional[str]]] = {}

        # 4. Extract card data sequentially (ordering matters)
        image_tasks = []  # (img_obj, img_path) tuples for threaded saving

        for i, image in enumerate(sorted_images):
            card_data, card_name, img_save_task = process_card_image(
                i, image, ws, merged_lookup, args, output_folder, len(cards), seen_prints
            )
            if card_data:
                cards.append(card_data)
                if img_save_task:
                    image_tasks.append((img_save_task, card_name))

        # 5. Save images in parallel threads
        if image_tasks:
            saved_count = 0
            skipped_count = 0
            with ThreadPoolExecutor(max_workers=MAX_IMAGE_THREADS) as executor:
                futures = {
                    executor.submit(save_image_if_changed, img_obj, img_path): card_name
                    for (img_obj, img_path), card_name in image_tasks
                }
                for future in as_completed(futures):
                    card_name = futures[future]
                    was_saved = future.result()
                    if was_saved:
                        saved_count += 1
                    else:
                        skipped_count += 1
            logger.info(f"🖼️  Images: {saved_count} saved, {skipped_count} unchanged")

        # 6. Save JSON
        json_path = os.path.join(output_folder, CARD_DATA_FILENAME)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(cards, f, ensure_ascii=False, indent=2)

        logger.info(f"✅ Success! Saved to: {json_path}")
        logger.info(f"📦 Total Cards: {len(cards)}")
        return len(cards)

    except Exception as e:
        logger.error(f"❌ Critical Error reading {xlsx_path}: {e}")
        return 0


def main() -> None:
    """Main entry point for card dataset generation."""
    parser = argparse.ArgumentParser(description="Generate JSON from Excel with card data.")
    parser.add_argument("--no-image", action="store_true", help="Skip image extraction and saving.")
    args = parser.parse_args()

    # Find XLSX files
    xlsx_files = []
    if os.path.exists(INPUT_FOLDER):
        for root, dirs, files in os.walk(INPUT_FOLDER):
            for file in files:
                if file.endswith(".xlsx") and not file.startswith("~$"):
                    xlsx_files.append(os.path.join(root, file))

    if not xlsx_files:
        logger.error(f"ไม่พบไฟล์ .xlsx ในโฟลเดอร์ {INPUT_FOLDER}")
        return

    # Process XLSX files in parallel threads
    total_cards = 0
    if len(xlsx_files) == 1:
        # Single file: process directly
        total_cards = process_single_xlsx(xlsx_files[0], args)
    else:
        # Multiple files: process in parallel
        logger.info(f"\n🚀 Processing {len(xlsx_files)} files with {min(MAX_FILE_THREADS, len(xlsx_files))} threads...")
        with ThreadPoolExecutor(max_workers=MAX_FILE_THREADS) as executor:
            futures = {
                executor.submit(process_single_xlsx, xlsx_path, args): xlsx_path
                for xlsx_path in xlsx_files
            }
            for future in as_completed(futures):
                xlsx_path = futures[future]
                try:
                    count = future.result()
                    total_cards += count
                except Exception as e:
                    logger.error(f"❌ Error processing {xlsx_path}: {e}")

    logger.info(f"\n🎯 Total cards across all files: {total_cards}")

    # Generate the root index file after processing all sheets
    generate_assets_index(ASSETS_FOLDER)

if __name__ == "__main__":
    main()